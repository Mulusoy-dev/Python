import pandas as pd
from openpyxl import Workbook, load_workbook
import time
import os


# load_workbook() ile Excel dosyası seçilir. 

word_file = load_workbook("C:/Users/melih/Desktop/exchange.xlsx")
word_status = word_file.active                                              # Çalışma alanı için 'active' değişkeni kullanılıyor

# print(word_file.sheetnames)                                                 # Aktif çalışma sayfasının adını yazdırma. Çıktı: ['Sayfa2']
# print(word_status)


#os.startfile("C:/Users/melih/Desktop/exchange.xlsx")
file = open("C:/Users/melih/Desktop/exchange.xlsx","r")
reading = file.read()

while True:
    print(f"USD selling: {word_status.cell(2,2).value} upgrade time: {word_status.cell(2,4).value}")
    print(f"EUR selling: {word_status.cell(3,2).value} upgrade time: {word_status.cell(3,4).value}")
    print(f"GBP selling: {word_status.cell(4,2).value} upgrade time: {word_status.cell(4,4).value}")
    print(f"CHF selling: {word_status.cell(5,2).value} upgrade time: {word_status.cell(5,4).value}")
    print(f"CAD selling: {word_status.cell(6,2).value} upgrade time: {word_status.cell(6,4).value}")
    # word_file.save("C:/Users/melih/Desktop/exchange.xlsx")
    time.sleep(30)
    





